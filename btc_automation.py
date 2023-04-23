import os
import sys
from typing import Any
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime
from pandas import DataFrame

app_path = os.path.dirname(sys.executable)
now = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
app_path = os.path.join(app_path, rf'BTC - {now}.xlsx')


def save_to_excel(content_df: DataFrame):
    content_df.to_excel("BTC_data.xlsx", 'Raport', startrow=2, startcol=1)
    wb = load_workbook('BTC_data.xlsx')
    sheet = wb['Raport']
    sheet.delete_cols(2, 2)
    min_column = wb.active.min_column
    max_column = wb.active.max_column
    min_row = wb.active.min_row
    max_row = wb.active.max_row
    # clear_index(sheet, min_column, min_row, max_row)
    sheet['A1'] = f'Raport BTC - {now}'
    sheet['A2'] = f'Interwał miesięczny'
    sheet['A1'].font = Font('Arial', bold=True, size=20)
    sheet['A2'].font = Font('Arial', bold=True, size=12)
    wb.save(f'BTC raport - {now[:-9]}.xlsx')
    print(min_row, max_row, min_column, max_column)


def clear_index(sheet: Any, min_col, min_row, max_row):
    for ind in range(min_row + 1, max_row + 1):
        sheet[f'{get_column_letter(min_col)}{ind}'] = ''
